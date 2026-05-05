import logging
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("TemplateGenerator")

# Configuration Constants
EXCEL_PASSWORD = "SPARSH_VERIFY_2026"
GRAY_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
HEADER_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top", horizontal="left")


def generate_consultant_template(
    master_df: pd.DataFrame, state: str, phase: str, output_path: str
) -> bool:
    try:
        # 1. Filter Data
        search_phase = str(phase).strip().lower()
        filtered_df = master_df[
            (master_df["state_canonical"].str.lower() == str(state).strip().lower())
            & (
                master_df["rusa_phase"].astype(str).str.strip().str.lower()
                == search_phase
            )
        ].copy()

        if filtered_df.empty:
            logger.warning(f"No projects found for {state} - {phase}.")
            return False

        # 2. Setup Headers and Mappings
        # Expanded to include full Release/Utilized history[cite: 7]
        header_display_map = {
            "state_canonical": "State",
            "rusa_phase": "Phase",
            "component": "Component",
            "District": "District",
            "inst_name": "Institution Name",
            "mpr_total_appr": "MPR Total Appr.",
            "mpr_central_appr": "MPR Central Appr.",
            "mpr_state_appr": "MPR State Appr.",
            "mpr_total_rel": "MPR Total Released",
            "mpr_central_rel": "MPR Central Released",
            "mpr_state_rel": "MPR State Released",
            "mpr_total_util": "MPR Total Utilized",
            "mpr_central_util": "MPR Central Utilized",
            "mpr_state_util": "MPR State Utilized",
            "uc_total_appr": "UC Total Appr.",
            "uc_central_appr": "UC Central Appr.",
            "uc_state_appr": "UC State Appr.",
            "uc_total_rel": "UC Total Released",
            "uc_central_rel": "UC Central Released",
            "uc_state_rel": "UC State Released",
            "uc_total_util": "UC Total Utilized",
            "uc_central_util": "UC Central Utilized",
            "uc_state_util": "UC State Utilized",
        }

        # Source columns from master_df
        export_cols = [
            "state_canonical",
            "rusa_phase",
            "component",
            "District",
            "inst_name",
            "mpr_total_appr",
            "mpr_central_appr",
            "mpr_state_appr",
            "mpr_total_rel",
            "mpr_central_rel",
            "mpr_state_rel",
            "mpr_total_util",
            "mpr_central_util",
            "mpr_state_util",
        ]

        # Entry fields: Totals are now excluded from fillable list[cite: 7]
        uc_fillable_fields = [
            "uc_central_appr",
            "uc_state_appr",
            "uc_central_rel",
            "uc_state_rel",
            "uc_central_util",
            "uc_state_util",
        ]

        # All UC-related headers
        uc_all_fields = [
            "uc_total_appr",
            "uc_central_appr",
            "uc_state_appr",
            "uc_total_rel",
            "uc_central_rel",
            "uc_state_rel",
            "uc_total_util",
            "uc_central_util",
            "uc_state_util",
        ]

        export_df = filtered_df[export_cols].copy()
        for field in uc_all_fields:
            export_df[field] = 0.0

        export_df["project_id_key"] = filtered_df["project_id_key"]
        export_df = export_df.rename(columns=header_display_map)

        # 3. Write to Excel
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            export_df.to_excel(writer, index=False, sheet_name="UC_Entry_Form")
            worksheet = writer.sheets["UC_Entry_Form"]

            # Map display names back to column letters for formula building
            col_map = {
                col: get_column_letter(i + 1) for i, col in enumerate(export_df.columns)
            }

            for col_idx, col_name in enumerate(export_df.columns, 1):
                col_letter = get_column_letter(col_idx)

                # Determine if column is User-Fillable[cite: 7]
                # Note: "UC Total" columns are Gray/Locked because they contain formulas
                is_fillable = any(
                    header_display_map.get(f) == col_name for f in uc_fillable_fields
                )

                current_fill = GREEN_FILL if is_fillable else GRAY_FILL
                current_prot = Protection(locked=not is_fillable)

                # Header Styles
                header_cell = worksheet[f"{col_letter}1"]
                header_cell.fill = current_fill
                header_cell.font = HEADER_FONT
                header_cell.border = THIN_BORDER
                header_cell.alignment = WRAP_ALIGNMENT

                # Column Widths
                if col_name == "Institution Name":
                    worksheet.column_dimensions[col_letter].width = 35
                elif "MPR" in col_name or "UC" in col_name:
                    worksheet.column_dimensions[col_letter].width = 16
                else:
                    worksheet.column_dimensions[col_letter].width = 15

                # Process Rows
                for row_idx in range(2, len(export_df) + 2):
                    cell = worksheet[f"{col_letter}{row_idx}"]
                    cell.fill = current_fill
                    cell.protection = current_prot
                    cell.border = THIN_BORDER

                    # Inject Formulas for UC Totals
                    if col_name == "UC Total Appr.":
                        cell.value = f"={col_map['UC Central Appr.']}{row_idx}+{col_map['UC State Appr.']}{row_idx}"
                        
                    # FIX: Removed the '.' after 'Released' to match your header_display_map
                    elif col_name == "UC Total Released": 
                        cell.value = f"={col_map['UC Central Released']}{row_idx}+{col_map['UC State Released']}{row_idx}"
                        
                    elif col_name == "UC Total Utilized":
                        cell.value = f"={col_map['UC Central Utilized']}{row_idx}+{col_map['UC State Utilized']}{row_idx}"

                    # Numeric Formatting
                    if "MPR" in col_name or "UC" in col_name:
                        cell.number_format = "#,##0.00"

                if col_name == "project_id_key":
                    worksheet.column_dimensions[col_letter].hidden = True

            worksheet.protection.password = EXCEL_PASSWORD
            worksheet.protection.enable()

        logger.info(f"Enhanced Formula Template generated at {output_path}[cite: 7].")
        return True

    except Exception as e:
        logger.error(f"Failed to generate template: {e}")
        return False
